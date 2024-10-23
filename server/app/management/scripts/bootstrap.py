from django.contrib.auth.models import User
from django.db import models
import app.models as models
from openpyxl import load_workbook, Workbook


def load_sample_data(source="sample_data.json", delete=False):
    # TODO: This method will load sample data to the app (ie. sample users, items etc.)
    if delete == True:
        delete_sample_data()

    if source.endswith(".json"):
        load_sample_data_from_json(filename=source)
    elif source.endswith(".xlsx"):
        load_sample_data_from_excel(filename=source)
    else:
        raise Exception(
            "Invalid source file type. Please use either .json or .xlsx. Got filename: {source}"
        )


def load_sample_data_from_json(filename="sample_data.json"):
    raise ("Not Implemented: load_sample_data_from_json")


def load_sample_data_from_excel(filename="sample_data.xlsx"):
    wb = load_workbook(filename=filename)
    load_sample_data_from_workbook(wb)


def load_sample_data_from_workbook(wb):
    # Create Application objects
    sheet = wb["Application"]
    headers = [cell.value for cell in sheet[1]]
    temp_applications = []
    for _row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in _row):
            continue
        row = dict(zip(headers, _row))
        temp_application = models.Application(
            name=row["name"],
            description=row["description"] or "",
            category=row["category"] or "",
        )
        temp_applications.append(temp_application)

    models.Application.objects.bulk_create(temp_applications, ignore_conflicts=True)

    # Create Shortcut objects
    sheet = wb["Shortcut"]
    headers = [cell.value for cell in sheet[1]]
    temp_shortcuts = []
    for _row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in _row):
            continue
        row = dict(zip(headers, _row))
        application = models.Application.objects.get(name=row["application_name"])
        temp_shortcut = models.Shortcut(
            application=application,
            submodule=row["submodule"] or "",
            command=row["command"],
            command_id=row["command_id"],
            context=row["context"] or "",
            mac=row["mac"] or "",
            windows=row["windows"] or "",
            linux=row["linux"] or "",
            level=row["level"],
            description=row["description"] or "",
            application_command=row["application_command"] or "",
        )
        temp_shortcuts.append(temp_shortcut)

    # bulk create Shortcut objects, ignore conflicts
    models.Shortcut.objects.bulk_create(temp_shortcuts, ignore_conflicts=True)

    # Create UserShortcut objects from df_user_shortcut
    sheet = wb["UserShortcut"]
    headers = [cell.value for cell in sheet[1]]
    temp_user_shortcuts = []
    for _row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in _row):
            continue
        row = dict(zip(headers, _row))
        shortcut = models.Shortcut.objects.get(
            application__name=row["application_name"],
            submodule=row["submodule"] or "",
            command_id=row["command_id"],
        )
        temp_user_shortcut = models.UserShortcut(
            user=User.objects.get(username=row["username"]),
            shortcut=shortcut,
            user_mac=row["user_mac"] or "",
            user_windows=row["user_windows"] or "",
            user_linux=row["user_linux"] or "",
            status=row["status"],
        )
        temp_user_shortcuts.append(temp_user_shortcut)

    # Bulk create UserShortcut objects, ignore conflicts
    models.UserShortcut.objects.bulk_create(temp_user_shortcuts, ignore_conflicts=True)


def export_data_to_workbook():
    # Create workbook
    wb = Workbook()
    del wb["Sheet"]

    # Export Application objects
    wb.create_sheet("Application")
    sheet = wb["Application"]
    applications = models.Application.objects.all()
    headers = ["name", "description", "category"]
    sheet.append(headers)
    for app in applications:
        sheet.append([app.name, app.description, app.category])

    # Export Shortcut objects
    wb.create_sheet("Shortcut")
    sheet = wb["Shortcut"]
    shortcuts = models.Shortcut.objects.all()
    headers = [
        "application_name",
        "submodule",
        "command_id",
        "command",
        "mac",
        "windows",
        "linux",
        "description",
        "application_command",
        "level",
        "category",
        "context",
    ]
    sheet.append(headers)
    for shortcut in shortcuts:
        sheet.append(
            [
                shortcut.application.name,
                shortcut.submodule,
                shortcut.command_id,
                shortcut.command,
                shortcut.mac,
                shortcut.windows,
                shortcut.linux,
                shortcut.description,
                shortcut.application_command,
                shortcut.level,
                shortcut.category,
                shortcut.context,
            ]
        )

    # Export UserShortcut objects
    wb.create_sheet("UserShortcut")
    sheet = wb["UserShortcut"]
    user_shortcuts = models.UserShortcut.objects.prefetch_related(
        "user", "shortcut", "shortcut__application"
    ).all()
    headers = [
        "username",
        "application_name",
        "submodule",
        "command_id",
        "user_mac",
        "user_windows",
        "user_linux",
        "status",
    ]
    sheet.append(headers)
    for user_shortcut in user_shortcuts:
        sheet.append(
            [
                user_shortcut.user.username,
                user_shortcut.shortcut.application.name,
                user_shortcut.shortcut.submodule,
                user_shortcut.shortcut.command_id,
                user_shortcut.user_mac,
                user_shortcut.user_windows,
                user_shortcut.user_linux,
                user_shortcut.status,
            ]
        )

    return wb


def create_admin_user(username, email, password):
    admin_exists = models.User.objects.filter(username=username).exists()
    if not admin_exists:
        models.User.objects.create_superuser(username, email, password)


def delete_sample_data():
    models.Application.objects.all().delete()
    models.Shortcut.objects.all().delete()
    models.UserShortcut.objects.all().delete()
