from io import BytesIO
import pytest
import os, django
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook, Workbook
from app.models import Application, Shortcut, UserShortcut
from core.models import Config

from app.management.scripts.bootstrap import (
    load_sample_data_from_excel,
    export_data_to_workbook,
)

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mpa.settings")
django.setup()


@pytest.mark.django_db
class TestBulkUploadView:
    @pytest.fixture
    def url(self):
        return "/api/bulk/upload_excel"

    @pytest.fixture
    def filename(self, user1):
        # We implicitly create user1, which is referred in the file
        filename = "server/app/tests/files/test_data.xlsx"
        Config.objects.create(name="SOURCE_DATA_FILENAME", value=filename)
        return "server/app/tests/files/test_data.xlsx"

    def test_no_file(self, ops_client, url):
        res = ops_client.post(url, data={})
        assert res.status_code == 400

    def test_upload(self, ops_client, url, filename):
        initial_application_count = Application.objects.count()
        initial_shortcut_count = Shortcut.objects.count()
        initial_usershortcut_count = UserShortcut.objects.count()

        with open(filename, "rb") as f:
            file_content = f.read()
            test_file = SimpleUploadedFile(
                "test_file.xlsx", file_content, content_type="text/xlsx"
            )
            data = {"file": test_file}
            res = ops_client.post(url, data=data)
            assert res.status_code == 200
            assert Application.objects.count() > initial_application_count
            assert Shortcut.objects.count() > initial_shortcut_count
            assert UserShortcut.objects.count() > initial_usershortcut_count

    def test_load_from_source(self, ops_client, filename):
        initial_application_count = Application.objects.count()
        initial_shortcut_count = Shortcut.objects.count()
        initial_usershortcut_count = UserShortcut.objects.count()

        url = "/api/bulk/load_from_source"
        res = ops_client.post(url)

        assert res.status_code == 200
        assert Application.objects.count() > initial_application_count
        assert Shortcut.objects.count() > initial_shortcut_count
        assert UserShortcut.objects.count() > initial_usershortcut_count

    def test_export_data_to_workbook(self, filename):
        load_sample_data_from_excel(filename=filename, create_users=True)
        wb = export_data_to_workbook()
        assert wb.sheetnames == ["Application", "Shortcut", "UserShortcut"]

        # Ensure all sheets have data
        assert wb["Application"].max_row > 1
        assert wb["Shortcut"].max_row > 1
        assert wb["UserShortcut"].max_row > 1

        # Ensure all columns are present
        assert wb["Application"].max_column > 1
        assert wb["Shortcut"].max_column > 1
        assert wb["UserShortcut"].max_column > 1

    def test_export(self, ops_client, filename):
        load_sample_data_from_excel(filename=filename, create_users=True)
        url = "/api/bulk/export_excel"
        res = ops_client.get(url)

        assert res.status_code == 200
        assert (
            res["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        wb = load_workbook(BytesIO(res.content))
        assert wb.sheetnames == ["Application", "Shortcut", "UserShortcut"]
        assert wb["Application"].max_row > 1
        assert wb["Shortcut"].max_row > 1
        assert wb["UserShortcut"].max_row > 1
